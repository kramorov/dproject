# pneumatic_actuators/models/pa_torque.py
from django.db import models
from django.utils.translation import gettext_lazy as _
from django.db.models import Q
from django.db import transaction
from django.http import HttpResponse

from datetime import datetime
import logging

from pneumatic_actuators.models import PneumaticActuatorSpringsQty
from pneumatic_actuators.models.pa_body import PneumaticActuatorBody
from params.models import PneumaticAirSupplyPressure
from pneumatic_actuators.models.py_options_constants import SPRINGS_DA_DEFAULT_CODE, SAFETY_POSITION_NC_DEFAULT_CODE, \
    ACTUATOR_VARIETY_RP_DEFAULT_CODE, ACTUATOR_VARIETY_SY_DEFAULT_CODE, SAFETY_POSITION_NO_DEFAULT_CODE, \
    SPRINGS_SR_DEFAULT_CODE

logger = logging.getLogger(__name__)


class BodyThrustTorqueTable(models.Model):
    """ Значения момента или усилия при заданном количестве пружин и давлении питания
        pressure - давление питания, code='spring' - для только пружин
        spring_qty - количество пружин,  code='DA' - привод без пружин, двойного действия
        body - модель (типоразмер) привода
        BTO (Beginning Torque to Open): Начальный момент открытия (момент страгивания для открытия).
            Это значение крутящего момента, необходимое для вывода запорного элемента арматуры из положения полного
            закрытия. Часто это наибольшее требуемое усилие из-за прилипания или уплотнения затвора.
        RTO (Running Torque to Open): Промежуточный момент открытия (или рабочий момент открытия).
            Это крутящий момент, необходимый во время основного движения запорного элемента (обычно в диапазоне
            от 0° до 90°, исключая крайние положения).
        ETO (End Torque to Open): Конечный момент открытия. Это крутящий момент, необходимый для достижения
            полностью открытого положения.
        BTC (Beginning Torque to Close): Начальный момент закрытия (момент страгивания для закрытия).
        RTC (Running Torque to Close): Промежуточный момент закрытия.
        ETC (End Torque to Close): Конечный момент закрытия (момент посадки затвора в уплотнение).
        В таблице указываются значения для НЗ привода - BTO, RTO, ETO
        При импорте для приводов с spring_qty.code='DA' все значения bto, rto, eto устанавливаются одинаковыми
            для приводов ДА имеет смысл только одно значение при заланном давлении
        Для НО привода при печати в заголовке значения должны инвертироваться на BTC, RTC, ETC
        геттеры:
        get_min_max_pressure_list_for_body(body, min_pressure, max_pressure) -возвращает список
            справочника PneumaticAirSupplyPressure, значение которых меньше или равно min_pressure,
            и больше или равно max_pressure
        get_torque_thrust_values(body_list, pressure_list default=null, spring_qty default=null или точное значение справочника,
         ncno default='NO')
         body_list - список ссылок на модель PneumaticActuatorBody, может состоять из одного элемента
         если pressure_list =null, spring_qty =null то возвращает
         структуру: заголовок и матрицу
         заголовок состоит из двух строк
            Первая строка - найденное в таблице давление, отсортированное по возрастанию sorting_order. Послений столбец "Пружины"
            Вторая строка - для каждого давления из первой строки делаем 1, 2 или 3 столбца значений:
                1 столбец - если привод DA - указываем bto
                2 стобца - если привод SR  и его тип шестерня-рейка - указываем bto, eto
                3 стоблца - если привод SR  и его тип кулисный (skotch-yoke) - указываем bto, rto, eto
                    Тип привода определяем через поле body, далее - PneumaticActuatorBody.model_line,
                    далее - PneumaticActuatorModelLine.pneumatic_actuator_construction_variety
            Первый столбец данных - body.code
            Второй стобец данных - spring_qty.code
            дальнейшие столбцы содержат данные из полей bto, rto, eto этой модели, отобранные в соответствии с заголовками
            т.е. какие поля брать и в каком порядке выводить.
        find_body_for_pressure_thrust_or_torque(pressure_min, pressure_max, thrust_or_torque, tolerance, PneumaticActuatorVariety)
            если PneumaticActuatorVariety.code='DA' то  возвращает список моделей, у которых
                есть pressure меньше или равно pressure_min, и значение bto
        """
    body = models.ForeignKey(PneumaticActuatorBody, on_delete=models.SET_NULL,
                             null=True, blank=True,  # ← ДОБАВЬТЕ ЭТО
                             related_name='body_thrust_torque_table',
                             verbose_name=_("Модель"),
                             help_text=_("Модель корпуса привода"))
    pressure = models.ForeignKey(PneumaticAirSupplyPressure, on_delete=models.SET_NULL,
                                 null=True, blank=True,  # ← ДОБАВЬТЕ ЭТО
                                 related_name='body_thrust_torque_table',
                                 verbose_name=_("Давление"),
                                 help_text=_("Давление питания или spring - для пружин"))
    spring_qty = models.ForeignKey(PneumaticActuatorSpringsQty, on_delete=models.SET_NULL,
                                   null=True, blank=True,  # ← ДОБАВЬТЕ ЭТО
                                   related_name='body_thrust_torque_table',
                                   verbose_name=_("Пружин / DA"),
                                   help_text=_("Количество пружин или DA"))
    bto = models.DecimalField(max_digits=10, decimal_places=1, verbose_name=_("Момент/усилие BTO"),
                              help_text=_("BTO Момент/усилие страгивания для открытия"))
    rto = models.DecimalField(max_digits=10, decimal_places=1, verbose_name=_("Момент/усилие RTC(MID)"),
                              help_text=_("RTC Момент/усилие в среднем положении"))
    eto = models.DecimalField(max_digits=10, decimal_places=1, verbose_name=_("Момент/усилие BTC"),
                              help_text=_("ETO Конечный Момент/Усилие открытия"))

    class Meta:
        verbose_name = _("Таблица моментов/усилий пневмоприводов")
        verbose_name_plural = _("Таблица моментов/усилий пневмоприводов")

    def __str__(self):
        return f"Таблица моментов/усилий для {self.body.name}"

        # ==================== ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ ====================

    @classmethod
    def _get_base_queryset(cls, current_body, pressure_list=None, spring_qty_list=None):
        """Базовый QuerySet для всех форматов"""
        queryset = cls.objects.filter(body__in=[current_body])

        if pressure_list:
            queryset = queryset.filter(pressure__in=pressure_list)

        if spring_qty_list:
            queryset = queryset.filter(spring_qty__in=spring_qty_list)

        return queryset



    # @classmethod
    # def _spring_sort_key(cls, spring_code):
    #     """Ключ сортировки для пружин"""
    #     if spring_code == SPRINGS_DA_DEFAULT_CODE:
    #         return 0
    #     try:
    #         return int(spring_code)
    #     except ValueError:
    #         return 999



    # ==================== ОСНОВНОЙ МЕТОД ====================

    @classmethod
    def get_torque_thrust_values(cls, current_body, pressure_list=None,
                                 spring_qty_list=None, ncno_code=SAFETY_POSITION_NO_DEFAULT_CODE,
                                 construction_variety_code=ACTUATOR_VARIETY_RP_DEFAULT_CODE, da_sr_code=SPRINGS_SR_DEFAULT_CODE):
        """
        Основной метод получения данных таблицы моментов/усилий

        Args:
            current_body: список объектов PneumaticActuatorBody или их ID
            pressure_list: список объектов PneumaticAirSupplyPressure или их ID (опционально)
            spring_qty_list: список объектов PneumaticActuatorSpringsQty или их ID (опционально)
            ncno_code: 'NO' или 'NC' - тип привода
            construction_variety_code: код конструкции - шестерня-рейка или кулисный

        Returns:
            Dict или  в зависимости от формата
        """
        logger = logging.getLogger(__name__)
        print(f"get_torque_thrust_values {current_body}")
        try:

            # Базовый запрос
            queryset = cls._get_base_queryset(current_body, pressure_list, spring_qty_list)
            structured_data = cls._format_structured_simple(queryset, ncno=ncno_code,
                                                            construction_variety_code=construction_variety_code,
                                                            da_sr_code=da_sr_code)
            print(f"structured_data {structured_data}")
            return structured_data
        except Exception as e:
            logger.error(f"Error in get_torque_thrust_values: {e}", exc_info=True)
            return {
                'error': str(e),
                'format': format,
                'data': [],
                'metadata': {}
            }
    # ==================== ФОРМАТЫ ВЫВОДА ====================

    @classmethod
    def _empty_optimized_response(cls) :
        """Пустой ответ в оптимизированном формате"""
        return {
            'format' : 'optimized' ,
            'body' : None ,
            'data' : {'by_spring' : {}} ,
            'table_config' : {
                'visible_fields' : [] ,
                'pressure_order' : [] ,
                'spring_order' : [] ,
                'field_descriptions' : {} ,
                'pressure_info' : {} ,
                'spring_info' : {} ,
                'format' : {
                    'torque' : {'unit' : 'Нм' , 'precision' : 1} ,
                    'pressure' : {'spring' : 'SPRING' , 'default_template' : '{value} бар'}
                }
            } ,
            'ncno' : 'nc' ,
            'construction_variety' : 'RP' ,
            'count' : 0
        }

    @classmethod
    def _error_optimized_response(cls , error , ncno='nc') :
        """Ответ с ошибкой в оптимизированном формате"""
        return {
            'format' : 'optimized' ,
            'error' : str(error) ,
            'body' : None ,
            'data' : {'by_spring' : {}} ,
            'table_config' : {
                'visible_fields' : [] ,
                'pressure_order' : [] ,
                'spring_order' : [] ,
                'field_descriptions' : {} ,
                'pressure_info' : {} ,
                'spring_info' : {} ,
                'format' : {
                    'torque' : {'unit' : 'Нм' , 'precision' : 1} ,
                    'pressure' : {'spring' : 'SPRING' , 'default_template' : '{value} бар'}
                }
            } ,
            'ncno' : ncno ,
            'construction_variety' : 'RP' ,
            'count' : 0
        }

    @classmethod
    def _spring_sort_key(cls , spring_code) :
        """Ключ для сортировки кодов пружин"""
        try :
            # Пытаемся преобразовать в число
            return int(spring_code)
        except (ValueError , TypeError) :
            # Если не число, оставляем как строку
            return spring_code

    @classmethod
    def _get_torque_fields_for_construction(cls, construction_variety_code, spring_code, ncno='NO'):
        """
        Определяет, какие поля моментов нужны для данного типа конструкции
        Возвращает список кортежей (field_name, display_name)
        """

        # Для DA приводов - только BTO
        if spring_code == SPRINGS_DA_DEFAULT_CODE:
            return [('bto', 'BTO' if ncno == SAFETY_POSITION_NC_DEFAULT_CODE else 'BTC')]

        # Для SR приводов (пружинных)
        if construction_variety_code == ACTUATOR_VARIETY_RP_DEFAULT_CODE:  # шестерня-рейка
            return [
                ('bto', 'BTO' if ncno == SAFETY_POSITION_NC_DEFAULT_CODE else 'BTC'),
                ('eto', 'ETO' if ncno == SAFETY_POSITION_NC_DEFAULT_CODE else 'ETC')
            ]
        elif construction_variety_code == ACTUATOR_VARIETY_SY_DEFAULT_CODE:  # кулисный
            return [
                ('bto', 'BTO' if ncno == SAFETY_POSITION_NC_DEFAULT_CODE else 'BTC'),
                ('rto', 'RTO' if ncno == SAFETY_POSITION_NC_DEFAULT_CODE else 'RTC'),
                ('eto', 'ETO' if ncno == SAFETY_POSITION_NC_DEFAULT_CODE else 'ETC')
            ]
        else:
            # По умолчанию все три
            return [
                ('bto', 'BTO' if ncno == SAFETY_POSITION_NC_DEFAULT_CODE else 'BTC'),
                ('rto', 'RTO' if ncno == SAFETY_POSITION_NC_DEFAULT_CODE else 'RTC'),
                ('eto', 'ETO' if ncno == SAFETY_POSITION_NC_DEFAULT_CODE else 'ETC')
            ]

    @classmethod
    def _format_structured_simple(cls , queryset , ncno='nc' , construction_variety_code='RP', da_sr_code=SPRINGS_SR_DEFAULT_CODE) :
        """
        Упрощенная версия для одного корпуса
        """
        try :
            all_data = queryset.select_related('body' , 'pressure' , 'spring_qty').order_by(
                'spring_qty__sorting_order' , 'pressure__sorting_order'
            )

            if not all_data :
                return cls._empty_optimized_response()

            # Получаем информацию о корпусе из первого элемента
            first_item = all_data[0]
            body = {
                'id' : first_item.body.id ,
                'code' : first_item.body.code ,
                'name' : first_item.body.name ,
            }

            # Инициализируем структуры
            by_spring = {}
            visible_fields = set()
            pressure_info = {}
            spring_info = {}
            available_pressures = set()
            available_springs = set()
            # Определяем поля моментов
            torque_fields = cls._get_torque_fields_for_construction(
                construction_variety_code, da_sr_code, ncno)

            for item in all_data :
                if not item.body or not item.spring_qty or not item.pressure :
                    continue

                spring_code = item.spring_qty.code

                # Инициализируем данные для пружин
                if spring_code not in by_spring :
                    by_spring[spring_code] = {
                        'pressures' : {} ,
                        'meta' : {
                            'id' : item.spring_qty.id ,
                            'code' : spring_code ,
                            'name' : item.spring_qty.name
                        }
                    }
                    spring_info[spring_code] = {
                        'id' : item.spring_qty.id ,
                        'name' : item.spring_qty.name ,
                        'sorting_order' : getattr(item.spring_qty , 'sorting_order' , 0)
                    }
                    available_springs.add(spring_code)



                # Формируем значения для данного давления
                pressure_code = item.pressure.code
                pressure_values = {}

                for field_name , display_name in torque_fields :
                    value = getattr(item , field_name , None)
                    if value is not None :
                        try :
                            pressure_values[field_name] = float(value)
                            visible_fields.add(field_name)
                        except (TypeError , ValueError) :
                            pressure_values[field_name] = None

                by_spring[spring_code]['pressures'][pressure_code] = pressure_values

                # Сохраняем информацию о давлении
                if pressure_code not in pressure_info :
                    pressure_info[pressure_code] = {
                        'id' : item.pressure.id ,
                        'code' : pressure_code ,
                        'name' : str(item.pressure) ,
                        'sorting_order' : getattr(item.pressure , 'sorting_order' , 0)
                    }
                    available_pressures.add(pressure_code)

            # Сортируем
            sorted_springs = sorted(
                available_springs ,
                key=lambda x : spring_info[x].get('sorting_order' , cls._spring_sort_key(x))
            )

            sorted_pressures = sorted(
                available_pressures ,
                key=lambda x : pressure_info[x].get('sorting_order' , 0)
            )

            # Стандартный порядок полей
            field_order = ['bto' , 'eto' , 'rto' , 'to']
            sorted_fields = [f for f in field_order if f in visible_fields]

            # Сортируем by_spring
            sorted_by_spring = {}
            for spring_code in sorted_springs :
                spring_data = by_spring[spring_code]
                # Сортируем давления внутри пружины
                sorted_pressures_data = {}
                for pressure_code in sorted_pressures :
                    if pressure_code in spring_data['pressures'] :
                        sorted_pressures_data[pressure_code] = spring_data['pressures'][pressure_code]
                spring_data['pressures'] = sorted_pressures_data
                sorted_by_spring[spring_code] = spring_data

            return {
                'format' : 'optimized' ,
                'body' : body ,
                'data' : {
                    'by_spring' : sorted_by_spring
                } ,
                'table_config' : {
                    'visible_fields' : sorted_fields ,
                    'pressure_order' : sorted_pressures ,
                    'spring_order' : sorted_springs ,
                    'field_descriptions' : {
                        'bto' : 'BTO (Break to Open)' ,
                        'eto' : 'ETO (End to Open)' ,
                        'rto' : 'RTO (Return to Open)' ,
                        'to' : 'TO (Torque)'
                    } ,
                    'pressure_info' : pressure_info ,
                    'spring_info' : spring_info ,
                    'format' : {
                        'torque' : {
                            'unit' : 'Нм' ,
                            'precision' : 1 ,
                            'template' : '{value:.1f} {unit}'
                        } ,
                        'pressure' : {
                            'spring' : 'SPRING' ,
                            'default_template' : '{value} бар'
                        }
                    }
                } ,
                'ncno' : ncno ,
                'construction_variety' : construction_variety_code ,
                'count' : len(sorted_by_spring)
            }

        except Exception as e :
            logger.error(f"Error in _format_structured_simple: {e}")
            return cls._error_optimized_response(e , ncno)

    # Вспомогательные методы для работы с данными

    @classmethod
    def get_torque_value(cls , formatted_data , spring_code , pressure_code , field='bto') :
        """
        Быстро получить значение момента из отформатированных данных

        Args:
            formatted_data: Данные в оптимизированном формате
            spring_code: Код пружин (например, '12')
            pressure_code: Код давления (например, '5.5')
            field: Поле момента ('bto', 'eto', 'rto', 'to')

        Returns:
            Значение момента или None
        """
        if not formatted_data or 'data' not in formatted_data :
            return None

        by_spring = formatted_data['data'].get('by_spring' , {})
        if spring_code not in by_spring :
            return None

        pressures = by_spring[spring_code].get('pressures' , {})
        if pressure_code not in pressures :
            return None

        return pressures[pressure_code].get(field)

    @classmethod
    def compare_with_reference(cls , formatted_data , spring_code , pressure_code ,
                               field , reference_value , tolerance_percent=10) :
        """
        Сравнить значение с контрольным

        Returns:
            Dict с результатами сравнения
        """
        actual = cls.get_torque_value(formatted_data , spring_code , pressure_code , field)

        if actual is None :
            return {
                'status' : 'error' ,
                'message' : f'Нет данных для пружин {spring_code}, давление {pressure_code}, поле {field}'
            }

        tolerance = reference_value * (tolerance_percent / 100)
        difference = actual - reference_value
        percentage = (difference / reference_value) * 100 if reference_value != 0 else 0

        return {
            'status' : 'success' ,
            'actual' : actual ,
            'reference' : reference_value ,
            'difference' : difference ,
            'percentage' : percentage ,
            'is_within_tolerance' : abs(difference) <= tolerance ,
            'tolerance_percent' : tolerance_percent ,
            'tolerance_value' : tolerance
        }

    @classmethod
    def generate_table_data(cls , formatted_data) :
        """
        Генерирует данные для таблицы в формате, удобном для фронтенда

        Returns:
            {
                'headers': [...],
                'rows': [...],
                'body': {...}
            }
        """
        if not formatted_data or 'data' not in formatted_data :
            return {'headers' : [] , 'rows' : [] , 'body' : None}

        body = formatted_data.get('body')
        table_config = formatted_data.get('table_config' , {})
        by_spring = formatted_data['data'].get('by_spring' , {})

        # Формируем заголовки
        headers = [{'key' : 'spring_qty' , 'label' : 'Пружины' , 'type' : 'string'}]

        pressure_order = table_config.get('pressure_order' , [])
        field_order = table_config.get('visible_fields' , [])
        field_descriptions = table_config.get('field_descriptions' , {})
        pressure_info = table_config.get('pressure_info' , {})

        for pressure_code in pressure_order :
            pressure_data = pressure_info.get(pressure_code , {})
            pressure_name = pressure_data.get('name' , pressure_code)

            for field in field_order :
                field_label = field_descriptions.get(field , field.upper())
                headers.append({
                    'key' : f'{pressure_code}:{field}' ,
                    'label' : f'{pressure_name} {field_label}' ,
                    'type' : 'number' ,
                    'pressure_code' : pressure_code ,
                    'field' : field ,
                    'unit' : table_config.get('format' , {}).get('torque' , {}).get('unit' , 'Нм')
                })

        # Формируем строки
        rows = []
        spring_order = table_config.get('spring_order' , sorted(by_spring.keys()))

        for spring_code in spring_order :
            if spring_code not in by_spring :
                continue

            row = {'spring_qty' : spring_code}
            spring_data = by_spring[spring_code]

            for pressure_code in pressure_order :
                if pressure_code in spring_data['pressures'] :
                    pressure_values = spring_data['pressures'][pressure_code]
                    for field in field_order :
                        value = pressure_values.get(field)
                        row[f'{pressure_code}:{field}'] = value
                else :
                    # Если нет данных для этого давления, заполняем None
                    for field in field_order :
                        row[f'{pressure_code}:{field}'] = None

            rows.append(row)

        return {
            'headers' : headers ,
            'rows' : rows ,
            'body' : body ,
            'table_config' : table_config
        }

    @classmethod
    def format_for_markdown(cls , formatted_data , include_metadata=True) :
        """
        Форматирует данные для вывода в Markdown

        Returns:
            Строка в формате Markdown
        """
        table_data = cls.generate_table_data(formatted_data)

        if not table_data['headers'] :
            return "Нет данных для отображения"

        # Заголовок таблицы
        body = formatted_data.get('body' , {})
        markdown_lines = []

        if body and body.get('name') :
            markdown_lines.append(f"## {body.get('name')} ({body.get('code')})")
            markdown_lines.append("")

        # Заголовки таблицы
        headers = table_data['headers']
        header_labels = [h['label'] for h in headers]
        markdown_lines.append("| " + " | ".join(header_labels) + " |")

        # Разделитель
        separator = "|" + "|".join(["---"] * len(headers)) + "|"
        markdown_lines.append(separator)

        # Данные
        for row in table_data['rows'] :
            row_values = []
            for header in headers :
                key = header['key']
                value = row.get(key)

                if value is None :
                    row_values.append("-")
                elif isinstance(value , (int , float)) :
                    # Форматируем числа
                    precision = formatted_data.get('table_config' , {}).get('format' , {}).get('torque' , {}).get(
                        'precision' , 1)
                    row_values.append(f"{value:.{precision}f}")
                else :
                    row_values.append(str(value))

            markdown_lines.append("| " + " | ".join(row_values) + " |")

        # Метаданные
        if include_metadata :
            markdown_lines.append("")
            markdown_lines.append("**Метаданные:**")
            markdown_lines.append(
                f"- Единицы измерения: {formatted_data.get('table_config' , {}).get('format' , {}).get('torque' , {}).get('unit' , 'Нм')}")
            markdown_lines.append(f"- Тип привода: {formatted_data.get('construction_variety' , 'N/A')}")
            markdown_lines.append(f"- Безопасное положение: {formatted_data.get('ncno' , 'N/A')}")

        return "\n".join(markdown_lines)
    # ==================== ДОПОЛНИТЕЛЬНЫЕ МЕТОДЫ ====================

    @classmethod
    def find_body_for_pressure_thrust_or_torque(cls, pressure_min, pressure_max,
                                                thrust_or_torque, tolerance,
                                                pneumatic_actuator_variety):
        """
        Поиск подходящих корпусов по давлению и моменту/усилию
        """
        queryset = cls.objects.select_related('body', 'pressure', 'spring_qty')

        # Фильтрация по давлению
        if pressure_min is not None:
            # Нужно адаптировать под вашу модель PneumaticAirSupplyPressure
            pass

        # Для DA приводов
        if pneumatic_actuator_variety.code == SPRINGS_DA_DEFAULT_CODE:
            queryset = queryset.filter(
                spring_qty__code=SPRINGS_DA_DEFAULT_CODE,
                bto__gte=thrust_or_torque - tolerance,
                bto__lte=thrust_or_torque + tolerance
            )

        # TODO: Добавить логику для других типов приводов

        return queryset.values_list('body', flat=True).distinct()

    @classmethod
    def get_min_max_pressure_list_for_body(cls, body, min_pressure, max_pressure):
        """
        Возвращает список давлений в заданном диапазоне для корпуса
        """

        pressures = PneumaticAirSupplyPressure.objects.filter(
            body_thrust_torque_table__body=body
        ).distinct()

        if min_pressure is not None:
            # Адаптируйте под вашу логику сравнения давлений
            pass

        if max_pressure is not None:
            # Адаптируйте под вашу логику сравнения давлений
            pass

        return pressures

    @staticmethod
    def export_table_template(pressure_min=2.5, pressure_max=8.0, springs_min=5, springs_max=12, output_path=None):
        """
        Экспорт таблицы моментов/усилий в Excel файл
        """
        from params.models import PneumaticAirSupplyPressure
        from pneumatic_actuators.models import PneumaticActuatorSpringsQty
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        import logging

        logger = logging.getLogger(__name__)

        try:
            # Получаем давления в диапазоне
            pressures = PneumaticAirSupplyPressure.objects.filter(
                Q(pressure_bar__gte=pressure_min) & Q(pressure_bar__lte=pressure_max)
            ).order_by('sorting_order')
            # добавляем давление с кодом 'spring' если оно существует
            spring_pressure = PneumaticAirSupplyPressure.objects.filter(code='spring').first()
            if spring_pressure and spring_pressure not in pressures:
                pressures = list(pressures) + [spring_pressure]

            # Получаем количества пружин в диапазоне
            if hasattr(PneumaticActuatorSpringsQty, 'value'):
                springs = PneumaticActuatorSpringsQty.objects.filter(
                    Q(value__gte=springs_min) & Q(value__lte=springs_max)
                ).order_by('sorting_order')
            else:
                # Если поля value нет, берем все активные пружины
                springs = PneumaticActuatorSpringsQty.objects.filter(
                    is_active=True
                ).order_by('sorting_order')

            # Получаем все корпуса
            bodies = PneumaticActuatorBody.objects.filter(is_active=True).order_by('sorting_order')

            # ИСПРАВЛЕНИЕ: получаем ВСЕ существующие данные для фильтрации
            all_torque_data = BodyThrustTorqueTable.objects.filter(
                body__in=bodies,
                pressure__in=pressures,
                spring_qty__in=springs
            ).select_related('body', 'pressure', 'spring_qty')

            # Создаем workbook
            wb = Workbook()

            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Создаем основной лист с данными
            ws = wb.active
            ws.title = "Моменты_усилия"

            # ИСПРАВЛЕНИЕ: создаем заголовки таблицы с двумя строками
            headers_row1 = ['', '', '']  # Пустые для корпуса и пружины
            headers_row2 = ['Код корпуса', 'Название корпуса', 'Код пружины']

            # Добавляем столбцы для каждого давления (BTO, RTO, ETO для каждого давления)
            for pressure in pressures:
                # Первая строка: КОД давления объединенный на 3 столбца
                headers_row1.extend([pressure.code, pressure.code, pressure.code])
                # Вторая строка: BTO, RTO, ETO
                headers_row2.extend(['BTO', 'RTO', 'ETO'])

            # Записываем заголовки - первая строка
            for col_idx, header in enumerate(headers_row1, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Вторая строка заголовков
            for col_idx, header in enumerate(headers_row2, 1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # ИСПРАВЛЕНИЕ: заполняем таблицу всеми комбинациями (начинаем с 3 строки)
            row_idx = 3

            # Получаем давление с кодом 'spring' для данных пружин
            spring_pressure = PneumaticAirSupplyPressure.objects.filter(code='spring').first()

            for body in bodies:
                for spring in springs:
                    # Начало строки: код корпуса, название, код пружины
                    row_data = [body.code, body.name, spring.code]

                    # Для каждого давления добавляем BTO, RTO, ETO
                    for pressure in pressures:
                        # Ищем данные для этой комбинации
                        torque_data = all_torque_data.filter(
                            body=body,
                            pressure=pressure,
                            spring_qty=spring
                        ).first()

                        if torque_data:
                            # Если данные есть - заполняем
                            row_data.extend([
                                torque_data.bto if torque_data.bto else '',
                                torque_data.rto if torque_data.rto else '',
                                torque_data.eto if torque_data.eto else ''
                            ])
                        else:
                            # Если данных нет - пустые ячейки
                            row_data.extend(['', '', ''])

                    # ИСПРАВЛЕНИЕ: добавляем данные для давления 'spring'
                    if spring_pressure:
                        spring_torque_data = all_torque_data.filter(
                            body=body,
                            pressure=spring_pressure,
                            spring_qty=spring
                        ).first()

                        if spring_torque_data:
                            row_data.extend([
                                spring_torque_data.bto if spring_torque_data.bto else '',
                                spring_torque_data.rto if spring_torque_data.rto else '',
                                spring_torque_data.eto if spring_torque_data.eto else ''
                            ])
                        else:
                            row_data.extend(['', '', ''])
                    else:
                        row_data.extend(['', '', ''])

                    # Записываем строку
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                    row_idx += 1

            # Создаем лист с параметрами
            ws_params = wb.create_sheet(title="Параметры")

            # Заголовки параметров
            param_headers = ['Параметр', 'Значение']
            for col_idx, header in enumerate(param_headers, 1):
                cell = ws_params.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Данные параметров
            params_data = [
                ['Минимальное давление, бар', pressure_min],
                ['Максимальное давление, бар', pressure_max],
                ['Минимальное кол-во пружин', springs_min],
                ['Максимальное кол-во пружин', springs_max],
                ['Дата экспорта', datetime.now().strftime('%Y-%m-%d %H:%M')]
            ]

            for row_idx, row_data in enumerate(params_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws_params.cell(row=row_idx, column=col_idx, value=value)

            # Создаем лист со справочниками
            ws_ref = wb.create_sheet(title="Справочники")

            # Заголовки справочников
            ref_headers = ['Тип', 'Код', 'Название', 'Описание']
            for col_idx, header in enumerate(ref_headers, 1):
                cell = ws_ref.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Заполняем справочник давлений
            row_idx = 2
            for pressure in pressures:
                ws_ref.cell(row=row_idx, column=1, value='Давление')
                ws_ref.cell(row=row_idx, column=2, value=pressure.code)
                ws_ref.cell(row=row_idx, column=3, value=pressure.name)
                ws_ref.cell(row=row_idx, column=4, value=pressure.description or '')
                row_idx += 1

            # Заполняем справочник пружин
            for spring in springs:
                ws_ref.cell(row=row_idx, column=1, value='Пружины')
                ws_ref.cell(row=row_idx, column=2, value=spring.code)
                ws_ref.cell(row=row_idx, column=3, value=spring.name)
                ws_ref.cell(row=row_idx, column=4, value=spring.description or '')
                row_idx += 1

            # Заполняем справочник корпусов
            for body in bodies:
                ws_ref.cell(row=row_idx, column=1, value='Корпус')
                ws_ref.cell(row=row_idx, column=2, value=body.code)
                ws_ref.cell(row=row_idx, column=3, value=body.name)
                ws_ref.cell(row=row_idx, column=4, value=body.description or '')
                row_idx += 1

            # Авто-ширина колонок для всех листов
            for worksheet in wb.worksheets:
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min((max_length + 2), 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем файл
            if output_path is None:
                output_path = f"torque_thrust_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

            wb.save(output_path)
            logger.info("Успешно экспортировано данных в: %s", output_path)
            return output_path

        except Exception as e:
            logger.error("Ошибка при экспорте в Excel: %s", str(e))
            raise

    @staticmethod
    @transaction.atomic
    def import_from_excel(excel_file_path):
        """
        Импорт данных моментов/усилий из Excel файла

        Args:
            excel_file_path: путь к файлу Excel

        Returns:
            tuple: (количество импортированных записей, список ошибок)
        """
        import pandas as pd
        from params.models import PneumaticAirSupplyPressure
        from pneumatic_actuators.models import PneumaticActuatorSpringsQty
        from pneumatic_actuators.models.pa_body import PneumaticActuatorBody

        try:
            # Читаем Excel файл
            df = pd.read_excel(excel_file_path, header=None)
            logger.info(f"Файл успешно прочитан. Размер: {df.shape}")
        except Exception as e:
            logger.error(f"Ошибка чтения файла: {str(e)}")
            return 0, [f"Ошибка чтения файла: {str(e)}"]

        imported_count = 0
        errors = []

        # Проверяем структуру файла
        if len(df.columns) < 4:
            error_msg = "Файл должен содержать минимум 4 столбца"
            logger.error(error_msg)
            return 0, [error_msg]

        # Получаем данные из столбцов
        body_codes_col = df.iloc[:, 0]  # Первый столбец - коды корпусов
        pressure_codes_row = df.iloc[0, 3:]  # Первая строка, начиная с 4-го столбца - коды давлений
        spring_codes_col = df.iloc[:, 2]  # Третий столбец - коды пружин
        parameter_types_col = df.iloc[1, 3:]  # Вторая строка, начиная с 4-го столбца - типы параметров

        logger.info(f"Первый столбец (корпуса): {body_codes_col.tolist()}")
        logger.info(f"Коды давлений из первой строки: {pressure_codes_row.tolist()}")
        logger.info(f"Третий столбец (пружины): {spring_codes_col.tolist()}")
        logger.info(f"Типы параметров из второй строки: {parameter_types_col.tolist()}")

        # Собираем информацию о столбцах с данными
        columns_info = []
        current_pressure = None

        for col_idx in range(3, len(df.columns)):  # Начинаем с 4-го столбца
            pressure_code = df.iloc[0, col_idx]  # Давление из первой строки
            parameter_type = df.iloc[1, col_idx]  # Тип параметра из второй строки

            if pd.notna(pressure_code) and str(pressure_code).strip():
                current_pressure = str(pressure_code).strip()

            if current_pressure and pd.notna(parameter_type) and str(parameter_type).strip():
                column_info = {
                    'col_idx': col_idx,
                    'pressure_code': current_pressure,
                    'parameter_type': str(parameter_type).strip().upper()
                }
                columns_info.append(column_info)
                logger.debug(f"Столбец {col_idx}: давление={current_pressure}, параметр={parameter_type}")

        logger.info(f"Всего столбцов с данными: {len(columns_info)}")

        # Получаем уникальные коды пружин из третьего столбца (начиная с 3-й строки)
        spring_codes = set()
        for row_idx in range(2, len(spring_codes_col)):  # Начинаем с 3-й строки (индекс 2)
            spring_code = spring_codes_col.iloc[row_idx]
            if pd.notna(spring_code) and str(spring_code).strip():
                spring_codes.add(str(spring_code).strip())

        logger.info(f"Уникальные коды пружин: {spring_codes}")

        # Проверяем существование всех давлений
        pressure_codes = set(info['pressure_code'] for info in columns_info)
        logger.info(f"Найдены коды давлений: {pressure_codes}")

        existing_pressures = PneumaticAirSupplyPressure.objects.filter(code__in=pressure_codes)
        existing_pressure_codes = set(p.code for p in existing_pressures)
        logger.info(f"Существующие давления в БД: {existing_pressure_codes}")

        missing_pressures = pressure_codes - existing_pressure_codes
        if missing_pressures:
            error_msg = f"Не найдены давления с кодами: {', '.join(missing_pressures)}"
            logger.error(error_msg)
            errors.append(error_msg)

        # Проверяем существование всех пружин
        existing_springs = PneumaticActuatorSpringsQty.objects.filter(code__in=spring_codes)
        existing_spring_codes = set(s.code for s in existing_springs)
        logger.info(f"Существующие пружины в БД: {existing_spring_codes}")

        missing_springs = spring_codes - existing_spring_codes
        if missing_springs:
            error_msg = f"Не найдены пружины с кодами: {', '.join(missing_springs)}"
            logger.error(error_msg)
            errors.append(error_msg)

        # Логируем все пружины в базе для отладки
        all_springs = PneumaticActuatorSpringsQty.objects.filter(is_active=True).values('code', 'name')
        logger.info(f"Все активные пружины в БД: {list(all_springs)}")

        # Если есть ошибки с давлениями или пружинами - прерываем импорт
        if errors:
            logger.error(f"Импорт прерван из-за ошибок: {errors}")
            return 0, errors

        # Создаем словари для быстрого доступа
        pressure_dict = {p.code: p for p in existing_pressures}
        spring_dict = {s.code: s for s in existing_springs}

        logger.info(f"Словарь давлений: {list(pressure_dict.keys())}")
        logger.info(f"Словарь пружин: {list(spring_dict.keys())}")

        # Обрабатываем данные начиная с 3-й строки
        logger.info(f"Начинаем обработку данных с {len(df) - 2} строк")

        for row_idx in range(2, len(df)):  # Начинаем с 3-й строки
            row = df.iloc[row_idx]
            logger.debug(f"Обработка строки {row_idx + 1}: {row.tolist()}")

            # Получаем код корпуса из первого столбца
            body_code = row.iloc[0] if len(row) > 0 else None
            # Получаем код пружины из третьего столбца
            spring_code = row.iloc[2] if len(row) > 2 else None

            if not body_code or pd.isna(body_code) or not spring_code or pd.isna(spring_code):
                logger.debug(f"Строка {row_idx + 1}: пропущена (пустой код корпуса или пружины)")
                continue

            body_code = str(body_code).strip()
            spring_code = str(spring_code).strip()
            logger.debug(f"Строка {row_idx + 1}: код корпуса '{body_code}', пружина '{spring_code}'")

            # Ищем корпус
            try:
                body = PneumaticActuatorBody.objects.get(code=body_code, is_active=True)
                logger.debug(f"Найден корпус: {body.name} (ID: {body.id})")
            except PneumaticActuatorBody.DoesNotExist:
                error_msg = f"Строка {row_idx + 1}: корпус с кодом '{body_code}' не найден"
                logger.error(error_msg)
                errors.append(error_msg)
                continue
            except PneumaticActuatorBody.MultipleObjectsReturned:
                error_msg = f"Строка {row_idx + 1}: найдено несколько корпусов с кодом '{body_code}'"
                logger.error(error_msg)
                errors.append(error_msg)
                continue

            # Получаем объект пружины
            spring = spring_dict.get(spring_code)
            if not spring:
                error_msg = f"Строка {row_idx + 1}: пружина с кодом '{spring_code}' не найдена"
                logger.error(error_msg)
                errors.append(error_msg)
                continue

            # Обрабатываем данные для каждого столбца
            for col_info in columns_info:
                col_idx = col_info['col_idx']

                if col_idx >= len(row):
                    logger.debug(f"Столбец {col_idx} выходит за пределы строки")
                    continue

                value = row.iloc[col_idx]
                logger.debug(f"Столбец {col_idx}: значение '{value}'")

                # Пропускаем пустые значения
                if pd.isna(value) or value == '':
                    logger.debug(f"Столбец {col_idx}: пропущено (пустое значение)")
                    continue

                try:
                    pressure = pressure_dict[col_info['pressure_code']]
                    parameter_type = col_info['parameter_type']

                    logger.debug(
                        f"Обработка: давление={pressure.code}, пружина={spring.code}, параметр={parameter_type}")

                    # Преобразуем значение в число
                    try:
                        numeric_value = float(value)
                        logger.debug(f"Значение преобразовано в число: {numeric_value}")
                    except (ValueError, TypeError):
                        error_msg = f"Строка {row_idx + 1}, столбец {col_idx + 1}: значение '{value}' не является числом"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        continue

                    # Ищем существующую запись или создаем новую
                    torque_data, created = BodyThrustTorqueTable.objects.get_or_create(
                        body=body,
                        pressure=pressure,
                        spring_qty=spring,
                        defaults={
                            'bto': 0,
                            'rto': 0,
                            'eto': 0
                        }
                    )

                    action = "создана" if created else "обновлена"
                    logger.debug(
                        f"Запись {action}: body={body.code}, pressure={pressure.code}, spring={spring.code}")

                    # Обновляем соответствующее поле
                    if parameter_type == 'BTO':
                        torque_data.bto = numeric_value
                        logger.debug(f"BTO установлено: {numeric_value}")
                    elif parameter_type == 'RTO':
                        torque_data.rto = numeric_value
                        logger.debug(f"RTO установлено: {numeric_value}")
                    elif parameter_type == 'ETO':
                        torque_data.eto = numeric_value
                        logger.debug(f"ETO установлено: {numeric_value}")
                    else:
                        error_msg = f"Строка {row_idx + 1}, столбец {col_idx + 1}: неизвестный тип параметра '{parameter_type}'"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        continue

                    torque_data.save()
                    imported_count += 1
                    logger.debug(f"Запись сохранена. Всего импортировано: {imported_count}")

                except Exception as e:
                    error_msg = f"Строка {row_idx + 1}, столбец {col_idx + 1}: ошибка обработки - {str(e)}"
                    logger.error(error_msg, exc_info=True)
                    errors.append(error_msg)
                    continue

        logger.info(f"Импорт завершен. Импортировано записей: {imported_count}, ошибок: {len(errors)}")
        return imported_count, errors
