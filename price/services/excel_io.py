# price/services/excel_io.py
"""
Excel import/export for PriceDocument.

Export: GET → .xlsx with columns: id (SKU), code, name, price
Import: POST multipart → reads Excel, syncs by SKU id:
  - new items added with price from file (or 0 if missing)
  - existing items updated (price only)
  - unknown SKU ids → warning
"""
import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.utils.translation import gettext_lazy as _

from price.models.price_document import PriceDocumentItem

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
HEADER_FONT = Font(bold=True)
HEADERS = ["id (SKU)", "code", "name", "price"]


def export_document_to_excel(doc) -> BytesIO:
    """
    Build an .xlsx workbook from document items and return as BytesIO.
    Columns: id (SKU), code, name, price.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = doc.name[:31] if doc.name else "Export"

    # Header row
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    items = (
        doc.items.filter(is_active=True)
        .select_related("sku")
        .order_by("sorting_order")
    )

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.sku_id or "")
        ws.cell(row=row_idx, column=2, value=item.sku.code if item.sku else "")
        ws.cell(row=row_idx, column=3, value=item.sku.name if item.sku else "")
        ws.cell(row=row_idx, column=4, value=float(item.price))

    # Auto-width for columns
    for col in range(1, 5):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_document_from_excel(doc, file) -> dict:
    """
    Read Excel, sync by SKU 'id' column.

    Returns dict: {created: int, updated: int, errors: [str]}
    """
    from openpyxl import load_workbook

    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    # Read header row to find columns
    header_map = {}
    for col_idx, cell in enumerate(ws[1], 1):
        val = str(cell.value or "").strip().lower()
        if val in ("id (sku)", "id", "sku_id"):
            header_map["id"] = col_idx
        elif val in ("code", "код"):
            header_map["code"] = col_idx
        elif val in ("name", "название", "наименование", "name"):
            header_map["name"] = col_idx
        elif val in ("price", "цена", "price"):
            header_map["price"] = col_idx

    if "id" not in header_map:
        return {"created": 0, "updated": 0, "errors": ["Column 'id (SKU)' not found in header"]}

    created = 0
    updated = 0
    errors = []

    # Build index of existing items keyed by sku_id
    existing_items = {
        item.sku_id: item
        for item in doc.items.filter(is_active=True).select_related("sku")
    }

    # Read data rows
    processed_sku_ids = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row):
            continue

        sku_id = _cell_value(row, header_map, "id")
        if sku_id is None or (isinstance(sku_id, str) and not sku_id.strip()):
            continue

        try:
            sku_id = int(sku_id)
        except (ValueError, TypeError):
            errors.append(f"Invalid SKU id: {sku_id}")
            continue

        price_str = _cell_value(row, header_map, "price")
        try:
            price = float(price_str) if price_str is not None and str(price_str).strip() else 0.0
        except (ValueError, TypeError):
            price = 0.0

        processed_sku_ids.add(sku_id)

        if sku_id in existing_items:
            # Update existing
            item = existing_items[sku_id]
            if float(item.price) != price:
                item.price = price
                item.save(update_fields=["price"])
                updated += 1
        else:
            # Create new item
            try:
                PriceDocumentItem.objects.create(
                    document=doc,
                    sku_id=sku_id,
                    price=price,
                    price_variety=doc.default_price_variety,
                    currency=doc.default_currency,
                )
                created += 1
            except Exception as e:
                errors.append(f"SKU id={sku_id}: {str(e)}")

    return {"created": created, "updated": updated, "errors": errors}


def _cell_value(row, header_map: dict, key: str):
    """Safely get cell value by column index from header_map."""
    col = header_map.get(key)
    if col is None:
        return None
    try:
        return row[col - 1]
    except IndexError:
        return None
