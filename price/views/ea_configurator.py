# price/views/ea_configurator.py
"""
API для конфигуратора цен электроприводов.

GET  /admin/prices/ea-configurator/options/?power_supply_id=X
     → доступные model_line_item + опции (encoding из through-моделей)
POST /admin/prices/ea-configurator/create/
     → создать документ + строки
GET  /admin/prices/ea-configurator/documents/
     → список документов конфигуратора
GET  /admin/prices/ea-configurator/documents/{id}/
     → документ со строками (матрица)
"""

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from project_customers.permissions import SectionAccessPermission

import logging
import pandas as pd
import numpy as np
from io import BytesIO
from django.db import transaction
from django.http import HttpResponse

logger = logging.getLogger(__name__)

from price.models import EAPriceDocument, EAPriceConstructor
from electric_actuators.models import (
    ElectricActuatorModelLineItem,
    ElectricPowerSupplyOption,
    ElectricActuatorConstructor,
    ElectricWaySwitchesOption,
)


class EaPowerSuppliesView(APIView):
    """Список всех доступных напряжений питания."""

    def get(self, request):
        supplies = ElectricPowerSupplyOption.objects.filter(is_active=True).select_related('power_supply')
        data = [{
            'id': ps.id,
            'name': str(ps),
            'encoding': ps.encoding,
        } for ps in supplies]
        return Response(data)


class EaConfiguratorOptionsView(APIView):
    """Доступные модели и опции для выбранного напряжения."""

    def get(self, request):
        ps_id = request.query_params.get('power_supply_id')
        logger.debug('EaConfiguratorOptions GET power_supply_id=%r', ps_id)
        if not ps_id:
            return Response({'error': 'power_supply_id required'}, status=400)

        try:
            ps = ElectricPowerSupplyOption.objects.get(id=ps_id)
        except ElectricPowerSupplyOption.DoesNotExist:
            logger.warning('EaConfiguratorOptions power_supply not found for id=%r', ps_id)
            return Response({'error': 'power_supply not found'}, status=404)

        # Все модели серии, у которых есть это напряжение (по справочнику PowerSupplies)
        ml_id = ps.model_line_item.model_line_id
        model_items = ElectricActuatorModelLineItem.objects.filter(
            model_line_id=ml_id,
            model_line_item_power_supply_option__power_supply_id=ps.power_supply_id,
            is_active=True,
        ).select_related('model_line', 'body').order_by('sorting_order').distinct()

        # Pre-fetch through-models for this series + voltage (for dependent options)
        all_ps = ElectricPowerSupplyOption.objects.filter(
            model_line_item__model_line_id=ml_id,
            power_supply_id=ps.power_supply_id,
            is_active=True,
        ).select_related('model_line_item')
        ps_by_mli = {p.model_line_item_id: p for p in all_ps}

        # Pre-fetch WaySwitches for model items of this series
        mli_ids = [item.id for item in model_items]
        way_by_mli = {}
        if mli_ids:
            for ws in ElectricWaySwitchesOption.objects.filter(
                model_line_item_id__in=mli_ids, is_active=True
            ).select_related('way_switches_option'):
                way_by_mli.setdefault(ws.model_line_item_id, []).append(ws)

        result = {
            'power_supply': {'id': ps.id, 'name': str(ps), 'encoding': ps.encoding},
            'model_items': [],
        }

        for item in model_items:
            item_ps = ps_by_mli.get(item.id)
            if not item_ps:
                continue
            temp = ElectricActuatorConstructor(
                selected_model_line_item=item,
                selected_power_supply=item_ps,
            )
            options = temp.get_available_options()

            # Преобразуем: для каждой группы опций берём encoding + name
            option_groups = []
            for key, items_list in options.items():
                if key == 'power_supply_options':
                    continue
                group = {
                    'field': key.replace('_options', ''),
                    'label': key,
                    'items': items_list,  # [{id, option_id, encoding, name, is_default}, ...]
                }
                option_groups.append(group)

            # WaySwitches — опции путевых выключателей
            ws_list = way_by_mli.get(item.id, [])
            if ws_list:
                option_groups.append({
                    'field': 'way_switches',
                    'label': 'way_switches',
                    'items': [{
                        'id': ws.id,
                        'option_id': ws.way_switches_option_id,
                        'encoding': ws.encoding or '',
                        'name': ws.way_switches_option.name,
                        'is_default': ws.is_default,
                    } for ws in ws_list],
                })

            result['model_items'].append({
                'id': item.id,
                'name': item.name,
                'code': item.code,
                'model_line_code': item.model_line.code,
                'torque_min': float(item.torque_min) if item.torque_min else None,
                'torque_max': float(item.torque_max) if item.torque_max else None,
                'rotation_speed': float(item.rotation_speed) if item.rotation_speed else None,
                'option_groups': option_groups,
            })

        return Response(result)


class EaConfiguratorDocumentView(APIView):
    """CRUD для документов конфигуратора."""
    permission_classes = [SectionAccessPermission]
    required_section = 'admin_section'

    def get(self, request, doc_id=None):
        if doc_id:
            url_name = request.resolver_match.url_name if request.resolver_match else ''
            if 'export' in url_name:
                return self._export(request, doc_id)
            if 'print' in url_name:
                return self._print_doc(request, doc_id)
            if 'fill' in url_name:
                return self._fill_prices(request, doc_id)
            return self._get_detail(doc_id)
        return self._list(request)

    def _list(self, request):
        qs = EAPriceDocument.objects.filter(is_active=True).exclude(status=EAPriceDocument.Status.DELETED).order_by('-document_date').select_related('price_variety', 'currency', 'power_supply', 'model_line')
        data = [{
            'id': d.id,
            'name': d.name,
            'document_date': str(d.document_date),
            'price_variety': {'id': d.price_variety_id, 'name': d.price_variety.name} if d.price_variety else None,
            'currency': {'id': d.currency_id, 'code': d.currency.code} if d.currency else None,
            'status': d.status,
            'status_label': d.get_status_display(),
            'power_supply': {'id': d.power_supply_id, 'name': str(d.power_supply)},
            'model_line': {'id': d.model_line_id, 'name': str(d.model_line)} if d.model_line else None,
            'rows_count': d.rows.count(),
        } for d in qs]
        return Response(data)

    def _get_detail(self, doc_id):
        try:
            doc = EAPriceDocument.objects.select_related('model_line', 'price_variety', 'currency', 'power_supply').get(id=doc_id)
        except EAPriceDocument.DoesNotExist:
            return Response({'error': 'not found'}, status=404)

        rows = EAPriceConstructor.objects.filter(document=doc).select_related('model_line_item').order_by('model_line_item', 'option_field', 'option_id')

        # Группируем строки по model_line_item
        by_model = {}
        for row in rows:
            mli_id = row.model_line_item_id
            if mli_id not in by_model:
                by_model[mli_id] = {
                    'model_line_item': {
                        'id': mli_id,
                        'name': row.model_line_item.name,
                        'code': row.model_line_item.code,
                    },
                    'base_price': None,
                    'options': {},
                }
            if row.option_field == 'base':
                by_model[mli_id]['base_price'] = float(row.surcharge)
            else:
                by_model[mli_id]['options'][f"{row.option_field}_{row.option_id}"] = float(row.surcharge)

        return Response({
            'id': doc.id,
            'name': doc.name,
            'document_date': str(doc.document_date),
            'price_variety': {'id': doc.price_variety_id, 'name': doc.price_variety.name} if doc.price_variety else None,
            'currency': {'id': doc.currency_id, 'code': doc.currency.code} if doc.currency else None,
            'status': doc.status,
            'status_label': doc.get_status_display(),
            'model_line': {'id': doc.model_line_id, 'name': doc.model_line.name, 'code': doc.model_line.code} if doc.model_line else None,
            'power_supply': {'id': doc.power_supply_id, 'name': str(doc.power_supply)},
            'rows': list(by_model.values()),
        })

    def post(self, request, doc_id=None):
        """POST /create/ → создать, /documents/{id}/post/ → провести, /unpost/ → отменить, /import/ → импорт.
           POST /documents/{id}/ → обновить (сохранить) существующий документ."""
        if doc_id is None:
            return self._create(request)
        url_name = request.resolver_match.url_name if request.resolver_match else ''
        if 'unpost' in url_name:
            return self._unpost(request, doc_id)
        if 'import' in url_name:
            return self._import(request, doc_id)
        if 'export' in url_name:
            return self._export(request, doc_id)
        if 'print' in url_name:
            return self._print_doc(request, doc_id)
        if 'post' in url_name:
            return self._conduct(request, doc_id)
        return self._update(request, doc_id)

    def _create(self, request):
        """Создать документ + строки из матрицы."""
        name = request.data.get('name', 'Конфигуратор цен')
        price_variety_id = request.data.get('price_variety_id')
        currency_id = request.data.get('currency_id')
        ps_id = request.data.get('power_supply_id')
        rows_data = request.data.get('rows', [])

        if not ps_id:
            return Response({'error': 'power_supply_id required'}, status=400)

        try:
            ps = ElectricPowerSupplyOption.objects.get(id=ps_id)
        except ElectricPowerSupplyOption.DoesNotExist:
            return Response({'error': 'power_supply not found'}, status=404)

        doc = EAPriceDocument.objects.create(
            name=name,
            price_variety_id=price_variety_id,
            currency_id=currency_id,
            model_line_id=request.data.get('model_line_id'),
            power_supply=ps,
            status=EAPriceDocument.Status.DRAFT,
        )

        if rows_data:
            self._create_rows(doc, rows_data, currency_id, price_variety_id)
        elif doc.model_line_id:
            # Auto-generate base rows for all models in this series+voltage
            model_items = ElectricActuatorModelLineItem.objects.filter(
                model_line_id=doc.model_line_id,
                model_line_item_power_supply_option__power_supply_id=ps.power_supply_id,
                is_active=True,
            ).distinct()
            for item in model_items:
                EAPriceConstructor.objects.create(
                    document=doc, model_line_item=item, power_supply=ps,
                    option_field='base', option_id=None, surcharge=0,
                    currency_id=currency_id, price_variety_id=price_variety_id,
                    is_active=False,
                )
        return Response({'id': doc.id, 'name': doc.name}, status=201)

    @transaction.atomic
    def _update(self, request, doc_id):
        """Обновить существующий документ (сохранить строки)."""
        try:
            doc = EAPriceDocument.objects.get(id=doc_id)
        except EAPriceDocument.DoesNotExist:
            return Response({'error': 'not found'}, status=404)

        if doc.status != EAPriceDocument.Status.DRAFT:
            return Response({'error': 'Only draft documents can be updated'}, status=400)

        doc.name = request.data.get('name', doc.name)
        doc.save(update_fields=['name', 'updated_at'])

        rows_data = request.data.get('rows')
        if rows_data is not None:
            currency_id = request.data.get('currency_id') or doc.currency_id
            price_variety_id = request.data.get('price_variety_id') or doc.price_variety_id
            EAPriceConstructor.objects.filter(document=doc).delete()
            self._create_rows(doc, rows_data, currency_id, price_variety_id)

        return Response({'id': doc.id, 'name': doc.name})

    def _conduct(self, request, doc_id):
        """Провести документ. Если уже проведён — отмена + пересоздание строк."""
        try:
            doc = EAPriceDocument.objects.get(id=doc_id)
        except EAPriceDocument.DoesNotExist:
            return Response({'error': 'not found'}, status=404)

        rows_data = request.data.get('rows')
        currency_id = request.data.get('currency_id')
        price_variety_id = request.data.get('price_variety_id')

        if rows_data and (not currency_id or not price_variety_id):
            return Response({'error': 'currency_id and price_variety_id required with rows'}, status=400)

        # Если уже проведён — отменить, пересоздать строки
        if doc.status == EAPriceDocument.Status.POSTED:
            doc.unpost()
            if rows_data:
                EAPriceConstructor.objects.filter(document=doc).delete()
                self._create_rows(doc, rows_data, currency_id, price_variety_id)

        # Если черновик и есть новые данные — пересоздать строки
        elif rows_data and doc.status == EAPriceDocument.Status.DRAFT:
            EAPriceConstructor.objects.filter(document=doc).delete()
            self._create_rows(doc, rows_data, currency_id, price_variety_id)

        doc.post()
        return Response({'ok': True, 'status': doc.status, 'status_label': doc.get_status_display()})

    def _unpost(self, request, doc_id):
        """Отменить проведение документа."""
        try:
            doc = EAPriceDocument.objects.get(id=doc_id)
        except EAPriceDocument.DoesNotExist:
            return Response({'error': 'not found'}, status=404)

        if doc.status != EAPriceDocument.Status.POSTED:
            return Response({'error': 'Документ не проведён'}, status=400)

        doc.unpost()
        return Response({'ok': True, 'status': doc.status, 'status_label': doc.get_status_display()})

    def _create_rows(self, doc, rows_data, currency_id, price_variety_id):
        """Создать строки EAPriceConstructor для документа."""
        ps = doc.power_supply
        for row in rows_data:
            mli_id = row.get('model_line_item_id')
            base_price = row.get('base_price', 0)
            if not mli_id:
                continue
            EAPriceConstructor.objects.create(
                document=doc, model_line_item_id=mli_id, power_supply=ps,
                option_field='base', option_id=None, surcharge=base_price,
                currency_id=currency_id, price_variety_id=price_variety_id,
                is_active=False,
            )
            for key, price in row.get('options', {}).items():
                if not price:
                    continue
                parts = key.rsplit('_', 1)
                if len(parts) != 2:
                    continue
                opt_field, opt_id = parts[0], parts[1]
                try:
                    opt_id_int = int(opt_id)
                except ValueError:
                    continue
                EAPriceConstructor.objects.create(
                    document=doc, model_line_item_id=mli_id, power_supply=ps,
                    option_field=opt_field, option_id=opt_id_int, surcharge=price,
                    currency_id=currency_id, price_variety_id=price_variety_id,
                    is_active=False,
                )

    def _get_option_columns(self, doc):
        """Get all option column definitions + per-model availability + label→mli_id.
        Returns {columns: [{key, label}], availability: {mli_id: set(key)}, label_to_mli: {str: int}}"""
        if not doc.power_supply:
            return {'columns': [], 'availability': {}, 'label_to_mli': {}}
        ps = doc.power_supply
        model_items = ElectricActuatorModelLineItem.objects.filter(
            model_line_id=doc.model_line_id,
            model_line_item_power_supply_option__power_supply_id=ps.power_supply_id,
            is_active=True,
        ).select_related('model_line').distinct()
        cols = []
        seen = set()
        availability = {}
        ps_enc = ps.encoding or ''
        label_to_mli = {}
        # Pre-fetch power supply options одним запросом
        mli_ids = [item.id for item in model_items]
        ps_by_mli = {}
        if mli_ids:
            for pso in ElectricPowerSupplyOption.objects.filter(
                model_line_item_id__in=mli_ids, power_supply_id=ps.power_supply_id, is_active=True
            ):
                ps_by_mli[pso.model_line_item_id] = pso
        for item in model_items:
            # label → mli_id (формат как в _build_matrix_df)
            label_to_mli[f"{item.model_line.code}{item.code}.{ps_enc}"] = item.id

            item_ps = ps_by_mli.get(item.id)
            if not item_ps:
                continue
            temp = ElectricActuatorConstructor(selected_model_line_item=item, selected_power_supply=item_ps)
            options = temp.get_available_options()
            item_keys = set()
            for key, items_list in options.items():
                if key == 'power_supply_options':
                    continue
                field = key.replace('_options', '')
                for opt in items_list:
                    if opt.get('is_default'):
                        continue
                    k = f"{field}_{opt['option_id']}"
                    item_keys.add(k)
                    if k not in seen:
                        seen.add(k)
                        cols.append({
                            'key': k,
                            'label': opt.get('encoding') or k,
                        })
            availability[item.id] = item_keys

        # WaySwitches — вне конструктора, привязаны напрямую к model_line_item
        mli_ids = [item.id for item in model_items]
        if mli_ids:
            way_qs = ElectricWaySwitchesOption.objects.filter(
                model_line_item_id__in=mli_ids, is_active=True
            ).select_related('way_switches_option')
            for ws in way_qs:
                k = f"way_switches_{ws.way_switches_option_id}"
                if k not in seen:
                    seen.add(k)
                    cols.append({
                        'key': k,
                        'label': ws.encoding or ws.way_switches_option.code or k,
                    })
                availability.setdefault(ws.model_line_item_id, set()).add(k)

        return {'columns': cols, 'availability': availability, 'label_to_mli': label_to_mli}

    def _build_matrix_df(self, doc):
        """Build pandas DataFrame from document rows + through-model option columns.
        NaN = unavailable option, 0 = available but not set, value = set."""
        col_data = self._get_option_columns(doc)
        col_defs = col_data['columns']
        availability = col_data['availability']

        base_rows = list(EAPriceConstructor.objects.filter(
            document=doc, option_field='base'
        ).select_related('model_line_item__model_line').order_by('model_line_item'))

        if not base_rows:
            return pd.DataFrame()

        ps_enc = doc.power_supply.encoding if doc.power_supply else ''

        # Build rows with NaN for all option columns
        rows = []
        for base in base_rows:
            mli = base.model_line_item
            label = f"{mli.model_line.code}{mli.code}.{ps_enc}"
            row = {'Модель': label, 'Базовая цена': float(base.surcharge)}
            for c in col_defs:
                row[c['label']] = np.nan
            rows.append(row)

        df = pd.DataFrame(rows)

        # Index DB options by model_line_item
        db_opts = EAPriceConstructor.objects.filter(document=doc).exclude(option_field='base')
        opt_by_mli = {}
        for opt in db_opts:
            opt_by_mli.setdefault(opt.model_line_item_id, {})[f"{opt.option_field}_{opt.option_id}"] = float(opt.surcharge)

        # Fill: available → 0 if not in DB, value if in DB; unavailable stays NaN
        for idx, base in enumerate(base_rows):
            mli_id = base.model_line_item_id
            avail = availability.get(mli_id, set())
            for c in col_defs:
                if c['key'] in avail:
                    df.at[idx, c['label']] = opt_by_mli.get(mli_id, {}).get(c['key'], 0.0)

        return df

    def _build_matrix_from_rows(self, doc, frontend_rows):
        """Build pandas DataFrame from frontend-provided rows (same format as matrix)."""
        col_data = self._get_option_columns(doc)
        col_defs = col_data['columns']
        availability = col_data['availability']
        ps_enc = doc.power_supply.encoding if doc.power_supply else ''

        rows = []
        for fr in frontend_rows:
            mli_id = fr.get('model_line_item_id')
            label = fr.get('label', str(mli_id))
            row = {'Модель': label, 'Базовая цена': float(fr.get('base_price', 0) or 0)}
            avail = availability.get(mli_id, set())
            for c in col_defs:
                if c['key'] in avail:
                    row[c['label']] = float((fr.get('options', {}).get(c['key'], 0)) or 0)
                else:
                    row[c['label']] = np.nan
            rows.append(row)

        if not rows:
            return pd.DataFrame()
        return pd.DataFrame(rows)

    def _export(self, request, doc_id):
        """POST /documents/{id}/export/ — скачать матрицу как Excel.
           Если передан {rows: [...]} — строит из них, иначе из БД."""
        try:
            doc = EAPriceDocument.objects.select_related('power_supply').get(id=doc_id)
        except EAPriceDocument.DoesNotExist:
            return Response({'error': 'not found'}, status=404)

        frontend_rows = request.data.get('rows') if request.method == 'POST' else None
        if frontend_rows:
            df = self._build_matrix_from_rows(doc, frontend_rows)
        else:
            df = self._build_matrix_df(doc)

        if df.empty:
            return Response({'error': 'no rows'}, status=404)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=(doc.name or 'EA_Configurator')[:31], index=False, na_rep='—')
            ws = writer.sheets[(doc.name or 'EA_Configurator')[:31]]
            from openpyxl.styles import Font, PatternFill
            hf = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            hfont = Font(bold=True)
            for col in range(1, len(df.columns) + 1):
                c = ws.cell(row=1, column=col)
                c.fill = hf
                c.font = hfont

        output.seek(0)
        safe_name = doc.name.replace(' ', '_').replace('/', '-')[:50] if doc.name else 'ea_config'
        resp = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="ea_config_{doc_id}_{safe_name}.xlsx"'
        return resp

    def _print_doc(self, request, doc_id):
        """POST /documents/{id}/print/ — HTML для печати.
           Если передан {rows: [...]} — строит из них, иначе из БД."""
        try:
            doc = EAPriceDocument.objects.select_related('model_line', 'price_variety', 'currency', 'power_supply').get(id=doc_id)
        except EAPriceDocument.DoesNotExist:
            return Response({'error': 'not found'}, status=404)

        frontend_rows = request.data.get('rows') if request.method == 'POST' else None
        if frontend_rows:
            df = self._build_matrix_from_rows(doc, frontend_rows)
        else:
            df = self._build_matrix_df(doc)

        table_html = df.to_html(index=False, na_rep='—', classes='mtx', border=0) if not df.empty else '<p>Нет данных</p>'

        html = f'''<!DOCTYPE html><html><head><meta charset="utf-8"><title>{doc.name}</title>
<style>body{{font-family:Arial;margin:20px;font-size:12px}}
h2{{margin:0 0 4px}}table.mtx{{border-collapse:collapse;width:100%;margin-top:10px}}
.mtx th,.mtx td{{border:1px solid #999;padding:4px 6px;text-align:left}}
.mtx th{{background:#d9e2f3}}.mtx td{{text-align:right}}</style></head><body>
<h2>{doc.name}</h2>
<p>Дата: {doc.document_date} | Тип цены: {doc.price_variety.name if doc.price_variety else '—'} | Валюта: {doc.currency.code if doc.currency else '—'}</p>
<p>Серия: {doc.model_line.name if doc.model_line else '—'} | Напряжение: {doc.power_supply}</p>
{table_html}</body></html>'''
        return Response({'html': html})

    def _import(self, request, doc_id):
        """POST /documents/{id}/import/ — загрузить Excel, разобрать на бэкенде,
        вернуть готовые строки: [{model_line_item_id, base_price, options: {key: price}}]."""
        try:
            doc = EAPriceDocument.objects.get(id=doc_id)
        except EAPriceDocument.DoesNotExist:
            return Response({'error': 'not found'}, status=404)

        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'error': 'No file'}, status=400)

        try:
            df = pd.read_excel(uploaded)
        except Exception:
            return Response({'error': 'Invalid file format'}, status=400)

        if df.empty or 'Модель' not in df.columns or 'Базовая цена' not in df.columns:
            return Response({'error': 'Invalid file format'}, status=400)

        col_data = self._get_option_columns(doc)
        label_to_key = {c['label']: c['key'] for c in col_data['columns']}
        label_to_mli = col_data['label_to_mli']

        option_cols = list(df.columns[2:])
        result_rows = []
        for _, excel_row in df.iterrows():
            model_label = str(excel_row.get('Модель', '')).strip()
            mli_id = label_to_mli.get(model_label)
            if not mli_id:
                continue

            options = {}
            for col_name in option_cols:
                val = excel_row[col_name]
                if pd.isna(val):
                    continue
                key = label_to_key.get(str(col_name))
                if key:
                    try:
                        options[key] = float(val)
                    except (ValueError, TypeError):
                        pass

            try:
                base_price = float(excel_row.get('Базовая цена', 0) or 0)
            except (ValueError, TypeError):
                base_price = 0.0

            result_rows.append({
                'model_line_item_id': mli_id,
                'base_price': base_price,
                'options': options,
            })

        return Response({'rows': result_rows})

    def _fill_prices(self, request, doc_id):
        """GET /documents/{id}/fill/ — действующие цены для заполнения матрицы."""
        try:
            doc = EAPriceDocument.objects.get(id=doc_id)
        except EAPriceDocument.DoesNotExist:
            return Response({'error': 'not found'}, status=404)

        if not doc.model_line_id or not doc.power_supply_id:
            return Response({'error': 'document has no series or power supply'}, status=400)

        active_rows = EAPriceConstructor.objects.filter(
            model_line_item__model_line_id=doc.model_line_id,
            power_supply=doc.power_supply,
            price_variety_id=doc.price_variety_id,
            currency_id=doc.currency_id,
            is_active=True,
        ).exclude(document=doc).select_related('model_line_item')

        by_model = {}
        for row in active_rows:
            mli_id = row.model_line_item_id
            if mli_id not in by_model:
                by_model[mli_id] = {
                    'model_line_item_id': mli_id,
                    'base_price': float(row.surcharge) if row.surcharge else 0.0,
                    'options': {},
                }
            if row.option_field == 'base':
                by_model[mli_id]['base_price'] = float(row.surcharge)
            else:
                by_model[mli_id]['options'][f"{row.option_field}_{row.option_id}"] = float(row.surcharge)

        return Response({'rows': list(by_model.values())})

    def delete(self, request, doc_id=None):
        """Мягкое удаление (mark_deleted)."""
        if not doc_id:
            return Response({'error': 'id required'}, status=400)
        try:
            doc = EAPriceDocument.objects.get(id=doc_id)
            doc.mark_deleted()
            return Response({'ok': True, 'status': doc.status})
        except EAPriceDocument.DoesNotExist:
            return Response({'error': 'not found'}, status=404)