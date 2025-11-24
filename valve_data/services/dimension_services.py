# valve_data/services/dimension_services.py
from django.db import transaction
import pandas as pd
import logging
from openpyxl import Workbook
from openpyxl.styles import Font , PatternFill

logger = logging.getLogger(__name__)

# Импортируем модели через общий модуль чтобы избежать циклических импортов
from valve_data.models import (
    ValveDimensionTable ,
    ValveDimensionData ,
    DimensionTableParameter ,
    DimensionTableDrawingItem ,
    WeightDimensionParameterVariety
)

# Импортируем модели из других приложений напрямую
from params.models import DnVariety , PnVariety


class DimensionDataService :
    """Сервис для работы с данными ВГХ"""

    @staticmethod
    def get_dimensions_display_data(dimension_table , dn_list=None , pn_list=None , export=False) :
        """
        Получить таблицу ВГХ в формате для отображения на экране с чертежами
        Использует новый универсальный геттер
        """
        logger.info("Начало формирования данных для отображения таблицы ВГХ")

        try :
            # Используем новый геттер из модели
            result = dimension_table.get_dimension_data(dn_list , pn_list , export)

            logger.info("Получены данные от геттера: %d изображений, %d матриц" ,
                        len(result['images']) , len(result['matrices']))

            # Формируем данные для отображения
            display_data = {
                'images' : result['images'] ,
                'matrices' : result['matrices'] ,
                'table_info' : {
                    'name' : dimension_table.name ,
                    'code' : dimension_table.code ,
                    'description' : dimension_table.description ,
                    'valve_brand' : dimension_table.valve_brand.name if dimension_table.valve_brand else None ,
                    'valve_variety' : dimension_table.valve_variety.symbolic_code if dimension_table.valve_variety else None
                } ,
                'export_mode' : export
            }

            logger.info("Успешно сформированы данные для отображения")
            return display_data

        except Exception as e :
            logger.error("Ошибка в get_display_data: %s" , str(e))
            logger.error("Трассировка:" , exc_info=True)
            raise

    @staticmethod
    def export_to_excel(dimension_table , dn_list=None , pn_list=None , output_path=None) :
        """
        Экспорт данных ВГХ в Excel файл для последующего импорта

        Args:
            dimension_table: ValveDimensionTable instance
            dn_list: список DN для экспорта
            pn_list: список PN для экспорта
            output_path: путь для сохранения файла

        Returns:
            str: путь к созданному файлу
        """
        try :
            # Получаем данные в экспортном формате (с кодами)
            result = dimension_table.get_dimension_data(dn_list , pn_list , export=True)

            if not result['matrices'] :
                raise ValueError("Нет данных для экспорта")

            # Создаем workbook
            wb = Workbook()

            # Стили для заголовков
            header_font = Font(bold=True , color="FFFFFF")
            header_fill = PatternFill(start_color="366092" , end_color="366092" , fill_type="solid")

            # Создаем основной лист с данными
            ws = wb.active
            ws.title = "Данные ВГХ"

            # Заголовки для экспорта
            headers = [
                'Обнозначение на схеме' ,
                'Код параметра' ,
                'PN код' ,
                'PN название'
            ]

            # Добавляем DN коды в заголовки
            if result['matrices'] :
                first_matrix = result['matrices'][0]['matrix']
                if first_matrix and len(first_matrix) > 0 :
                    # Пропускаем первые 4 колонки (legend, text_value, parameter_variety_code, pn)
                    dn_headers = first_matrix[0][4 :]
                    headers.extend(dn_headers)

            # Записываем заголовки
            for col_idx , header in enumerate(headers , 1) :
                cell = ws.cell(row=1 , column=col_idx , value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Записываем данные
            row_idx = 2
            for matrix_data in result['matrices'] :
                matrix = matrix_data['matrix']
                pn_code = matrix_data['pn']

                # Пропускаем строку заголовков матрицы и обрабатываем данные
                for row in matrix[1 :] :  # Пропускаем первую строку (заголовки)
                    logger.info(f"row_idx:{row_idx} row:{row}")
                    if len(row) >= 4 :  # Минимум 4 колонки должно быть
                        # Формируем строку для экспорта
                        import_row = [
                            row[0] ,  # legend -
                            row[1] ,  # PARAM_NAME - название параметра
                            row[2] ,  # parameter_variety_code - код параметра
                            pn_code ,  # PN код

                        ]

                        # Добавляем значения DN
                        import_row.extend(row[4 :])  # Значения DN

                        # Записываем строку
                        for col_idx , value in enumerate(import_row , 1) :
                            ws.cell(row=row_idx , column=col_idx , value=value)

                        row_idx += 1

            # Создаем лист со справочниками
            ws_ref = wb.create_sheet(title="Справочники")

            # Заголовки справочников
            ref_headers = ['Тип' , 'Код' , 'Название' , 'Описание']
            for col_idx , header in enumerate(ref_headers , 1) :
                cell = ws_ref.cell(row=1 , column=col_idx , value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Заполняем справочник DN
            row_idx = 2
            dns = DnVariety.objects.filter(is_active=True).order_by('sorting_order')
            for dn in dns :
                ws_ref.cell(row=row_idx , column=1 , value='DN')
                ws_ref.cell(row=row_idx , column=2 , value=dn.code)
                ws_ref.cell(row=row_idx , column=3 , value=dn.name)
                ws_ref.cell(row=row_idx , column=4 , value=dn.description or '')
                row_idx += 1

            # Заполняем справочник PN
            pns = PnVariety.objects.filter(is_active=True).order_by('sorting_order')
            for pn in pns :
                ws_ref.cell(row=row_idx , column=1 , value='PN')
                ws_ref.cell(row=row_idx , column=2 , value=pn.code)
                ws_ref.cell(row=row_idx , column=3 , value=pn.name)
                ws_ref.cell(row=row_idx , column=4 , value=pn.description or '')
                row_idx += 1

            # Авто-ширина колонок для всех листов
            for ws in wb.worksheets :
                for column in ws.columns :
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column :
                        try :
                            if len(str(cell.value)) > max_length :
                                max_length = len(str(cell.value))
                        except :
                            pass
                    adjusted_width = min((max_length + 2) , 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем файл
            if output_path is None :
                output_path = f"dimension_export_{dimension_table.code}.xlsx"

            wb.save(output_path)
            logger.info("Успешно экспортировано данных в: %s" , output_path)
            return output_path

        except Exception as e :
            logger.error("Ошибка при экспорте в Excel: %s" , str(e))
            raise

    @staticmethod
    @transaction.atomic
    def import_data_to_table(dimension_table , excel_file_path) :
        """
        Импорт данных ВГХ из Excel файла с полной перезаписью существующих данных

        Args:
            dimension_table: ValveDimensionTable instance
            excel_file_path: путь к файлу Excel

        Returns:
            tuple: (количество импортированных записей, список ошибок)
        """
        try :
            df = pd.read_excel(excel_file_path)
        except Exception as e :
            return 0 , [f"Ошибка чтения файла: {str(e)}"]

        imported_count = 0
        errors = []

        # ВРЕМЕННОЕ ХРАНИЛИЩЕ ДЛЯ НОВЫХ ДАННЫХ
        new_data_records = []

        # Используем enumerate для получения номера строки
        for row_index , (index , row) in enumerate(df.iterrows() , start=2) :
            try :
                # Получаем данные из фиксированных столбцов
                # Первый столбец - название параметра
                param_name = row.iloc[0] if len(row) > 0 else None
                # Второй столбец - код параметра
                param_code = row.iloc[1] if len(row) > 1 else None
                # Третий столбец - PN код
                pn_code = row.iloc[2] if len(row) > 2 else None

                # Проверяем обязательные поля
                if not param_name or pd.isna(param_name) :
                    errors.append(f"Строка {row_index}: отсутствует название параметра")
                    continue

                if not pn_code or pd.isna(pn_code) :
                    errors.append(f"Строка {row_index}: отсутствует PN код")
                    continue

                # Ищем параметр по коду используя геттер из модели
                param_variety = None
                if param_code and pd.notna(param_code) :
                    try :
                        param_variety = WeightDimensionParameterVariety.objects.get(
                            code=param_code , is_active=True
                        )
                    except WeightDimensionParameterVariety.DoesNotExist :
                        errors.append(f"Строка {row_index}: параметр с кодом '{param_code}' не найден")
                        continue

                # Создаем или получаем параметр таблицы
                table_param , created = DimensionTableParameter.objects.get_or_create(
                    dimension_table=dimension_table ,
                    name=param_name ,
                    defaults={
                        'parameter_variety' : param_variety ,
                        'sorting_order' : DimensionTableParameter.objects.filter(
                            dimension_table=dimension_table
                        ).count() + 1
                    }
                )

                # Если параметр уже существует, обновляем связь с системным параметром
                if not created and param_variety and not table_param.parameter_variety :
                    table_param.parameter_variety = param_variety
                    table_param.save()

                # Ищем PN используя геттер из модели
                pn = PnVariety.find_pn(str(pn_code))
                if not pn :
                    errors.append(f"Строка {row_index}: PN '{pn_code}' не найден")
                    continue

                # Обрабатываем значения DN начиная с 4-го столбца
                for col_index in range(4 , len(row)) :  # Начинаем с 4-го столбца (после PN названия)
                    # Получаем заголовок столбца из первой строки DataFrame
                    if col_index >= len(df.columns) :
                        break

                    col_header = str(df.columns[col_index])

                    # Пропускаем пустые заголовки
                    if not col_header or pd.isna(col_header) or col_header.strip() == '' :
                        break

                    # Получаем значение из текущей строки
                    value = row.iloc[col_index] if col_index < len(row) else None

                    # Ищем DN используя геттер из модели
                    dn = DnVariety.find_dn(col_header)
                    if not dn :
                        errors.append(f"Строка {row_index}: DN '{col_header}' не найден")
                        continue

                    # Обрабатываем значение
                    num_value = None
                    text_val = ""

                    if pd.notna(value) and value not in ['' , 'н/д' , 'N/A' , None] :
                        try :
                            # Пробуем преобразовать в число
                            num_value = float(value)
                        except (ValueError , TypeError) :
                            # Если не число, сохраняем как текст
                            text_val = str(value) if value is not None else ""

                    # СОЗДАЕМ ОБЪЕКТ ДЛЯ ВРЕМЕННОГО ХРАНЕНИЯ
                    new_data_records.append(ValveDimensionData(
                        dn=dn ,
                        pn=pn ,
                        parameter=table_param ,
                        value=num_value ,
                        text_value=text_val
                    ))

                    imported_count += 1
                    logger.debug("Подготовлена запись: %s, %s, %s" , dn.code , pn.code , param_name)

            except Exception as e :
                errors.append(f"Строка {row_index}: ошибка обработки - {str(e)}")
                logger.error("Ошибка в строке %d: %s" , row_index , str(e))
                continue

        # УДАЛЯЕМ СТАРЫЕ ДАННЫЕ ТОЛЬКО ЕСЛИ ЕСТЬ НОВЫЕ ДАННЫЕ
        if new_data_records and not errors :
            table_parameters = DimensionTableParameter.objects.filter(
                dimension_table=dimension_table
            )

            # Удаляем данные измерений
            deleted_count , _ = ValveDimensionData.objects.filter(
                parameter__in=table_parameters
            ).delete()

            logger.info("Удалено существующих записей: %d" , deleted_count)

            # СОЗДАЕМ НОВЫЕ ДАННЫЕ
            ValveDimensionData.objects.bulk_create(new_data_records)
            logger.info("Создано новых записей: %d" , len(new_data_records))

        elif errors :
            logger.warning("Импорт отменен из-за ошибок. Ошибок: %d" , len(errors))
            imported_count = 0  # Сбрасываем счетчик, так как данные не были сохранены

        logger.info("Импорт завершен. Импортировано записей: %d, ошибок: %d" ,
                    imported_count , len(errors))
        return imported_count , errors

    @staticmethod
    def create_empty_template(dimension_table , output_path=None) :
        """
        Создать пустой шаблон Excel для заполнения данных ВГХ
        """
        # Получаем все системные параметры
        parameters = WeightDimensionParameterVariety.objects.filter(
            is_active=True
        ).order_by('name')

        # Получаем все PN и DN
        pns = PnVariety.objects.all().order_by('sorting_order')
        dns = DnVariety.objects.all().order_by('sorting_order')

        # Создаем workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Шаблон ВГХ"

        # Стили для заголовков
        header_font = Font(bold=True , color="FFFFFF")
        header_fill = PatternFill(start_color="366092" , end_color="366092" , fill_type="solid")

        # Заголовки для импорта
        headers = [
                      'Название параметра' ,
                      'Код параметра' ,
                      'PN код' ,
                      'PN название'
                  ] + [f'DN{dn.code}' for dn in dns]

        for col , header in enumerate(headers , 1) :
            cell = ws.cell(row=1 , column=col , value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Заполняем данными
        row_num = 2

        # Заполняем параметры
        for param in parameters :
            # Название параметра
            ws.cell(row=row_num , column=1 , value=param.name)
            # Код параметра
            ws.cell(row=row_num , column=2 , value=param.code)
            row_num += 1

        # Авто-ширина колонок
        for column in ws.columns :
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column :
                try :
                    if len(str(cell.value)) > max_length :
                        max_length = len(str(cell.value))
                except :
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем файл
        if output_path is None :
            output_path = f"dimension_template_{dimension_table.code}.xlsx"

        wb.save(output_path)
        return output_path

    @staticmethod
    @transaction.atomic
    def clear_pn_data(dimension_table , pn_codes) :
        """
        Удалить все данные ВГХ для указанных PN

        Args:
            dimension_table: ValveDimensionTable instance
            pn_codes: список кодов PN для очистки

        Returns:
            int: количество удаленных записей
        """
        pns = PnVariety.objects.filter(code__in=pn_codes)
        table_parameters = DimensionTableParameter.objects.filter(
            dimension_table=dimension_table
        )

        deleted_count , _ = ValveDimensionData.objects.filter(
            pn__in=pns ,
            parameter__in=table_parameters
        ).delete()

        return deleted_count